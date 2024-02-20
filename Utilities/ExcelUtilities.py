import openpyxl
from openpyxl.styles import PatternFill

def getrowno(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def getcolumnno(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readdata(file,sheetname,rowno,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rowno,columnno).value

def Writedata(file,sheetname,rowno,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rowno,columnno).value=data
    workbook.save()